import json
import sys
import os
import glob
from openpyxl import load_workbook

def update_excel_bulk(excel_file, sheet_name):
    print(f"\n{'*'*115}")
    print(f"INITIALIZING BULK PROCESS")
    print(f"{'*'*115}")

    # 1. Validation
    if not os.path.exists(excel_file):
        print(f"Error: Excel file '{excel_file}' not found.")
        return

    # Find all .txt files (case-insensitive search)
    text_files = [f for f in os.listdir('.') if f.lower().endswith('.txt')]
    
    if not text_files:
        print("No text files (.txt) found in the current directory.")
        return
    
    print(f"Found {len(text_files)} text files: {', '.join(text_files)}")

    # 2. Load Excel Workbook
    try:
        wb = load_workbook(excel_file)
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return

    # Identify Column Indices
    header = [cell.value for cell in ws[1]]
    try:
        concat_col_idx = header.index('Concat Name') + 1
        target_col_idx = header.index('Target folder') + 1
    except ValueError:
        print("Error: Could not find 'Concat Name' or 'Target folder' headers.")
        return

    status_mapping_rules = {
        'Correct Info': 'Completed',
        'Needs Additional Information': 'Needs additional information',
        'Missing TS': 'Missing TS'
    }

    total_updates = 0

    # 3. Process each text file
    for text_file in text_files:
        print(f"\n>> STARTING FILE: {text_file}")
        print(f"{'-'*115}")
        print(f"{'PDF Name':<60} | {'Match':<10} | {'New Status'}")
        print(f"{'-'*115}")
        
        try:
            # Using utf-8-sig to handle files with BOM (common in Windows copies)
            with open(text_file, 'r', encoding='utf-8-sig') as f:
                json_data = json.load(f)
            
            status_map = {
                item['PDF Name']: status_mapping_rules.get(item['Status'], item['Status'])
                for item in json_data if 'PDF Name' in item
            }
        except Exception as e:
            print(f"FAILED TO READ {text_file}: {e}")
            continue

        file_updates = 0
        # Iterate through Excel rows
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=concat_col_idx).value
            if cell_val is None:
                continue
                
            concat_val = str(cell_val).strip()
            
            if concat_val in status_map:
                new_status = status_map[concat_val]
                ws.cell(row=row, column=target_col_idx).value = new_status
                
                display_name = (concat_val[:57] + "...") if len(concat_val) > 60 else concat_val
                print(f"{display_name:<60} | SUCCESS    | {new_status}")
                file_updates += 1
                total_updates += 1
                
        print(f"{'-'*115}")
        print(f"COMPLETED {text_file}: {file_updates} updates made.")

    # 4. Save Workbook once at the end
    try:
        wb.save(excel_file)
        print(f"\n{'='*115}")
        print(f"SUMMARY REPORT")
        print(f"{'='*115}")
        print(f"Total Text Files Processed: {len(text_files)}")
        print(f"Total Excel Rows Updated:   {total_updates}")
        print(f"Result Saved to:            {excel_file}")
    except PermissionError:
        print(f"\nPERMISSION ERROR: Please close '{excel_file}' and run again.")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python P2D_Bulk_Update.py Brio.xlsx \"20-Apr-2026\"")
    else:
        update_excel_bulk(sys.argv[1], sys.argv[2])